from openpyxl import Workbook, load_workbook

wb = load_workbook('testbook.xlsx')

# Load workshit
ws = wb.active
#print(ws['A1'].value)


sheets = wb.sheetnames
#print(sheets)

# Adding / appending rows
#!!! Нам нужно собрать лист из всех необходимых значений
#!!! и передать его в следующий метод. Метод заполняет
#!!! первый свободный ряд в активной странице экселя.
type = 'Labwork'
qty = 24
resp = 'Lena'

ws.append([type, qty, resp])


# Saving workshit
wb.save('testbook.xlsx')